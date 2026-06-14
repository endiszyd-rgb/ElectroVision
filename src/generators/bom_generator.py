"""BOM export: CSV, Excel with grouping and formatting."""
from pathlib import Path
from collections import defaultdict
from src.core.models.component import Component


class BOMGenerator:
    @staticmethod
    def group_components(components: list[Component]) -> list[dict]:
        groups: dict[str, list[Component]] = defaultdict(list)
        for comp in components:
            key = f"{comp.value}|{comp.footprint}"
            groups[key].append(comp)
        rows = []
        for comps in groups.values():
            c = comps[0]
            rows.append({
                "Reference": ", ".join(x.reference for x in comps),
                "Wartość":   c.value,
                "Typ":       c.component_type,
                "Footprint": c.footprint.split(":")[-1] if ":" in c.footprint else c.footprint,
                "Ilość":     len(comps),
                "Producent": c.manufacturer,
                "Nr kat.":   c.manufacturer_pn,
                "Datasheet": c.datasheet,
                "Opis":      c.description,
            })
        return sorted(rows, key=lambda r: r["Typ"])

    @staticmethod
    def to_csv(components: list[Component], path: str) -> None:
        import csv
        rows = BOMGenerator.group_components(components)
        if not rows:
            return
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
            writer.writeheader()
            writer.writerows(rows)

    @staticmethod
    def to_excel(components: list[Component], path: str) -> None:
        import pandas as pd
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        rows = BOMGenerator.group_components(components)
        if not rows:
            return

        df = pd.DataFrame(rows)
        df.to_excel(path, index=False, sheet_name="BOM")

        wb = openpyxl.load_workbook(path)
        ws = wb.active

        header_font   = Font(bold=True, color="FFFFFF", size=11)
        header_fill   = PatternFill("solid", fgColor="1A5C1A")
        alt_fill      = PatternFill("solid", fgColor="F0F8F0")
        thin_side     = Side(style="thin", color="CCCCCC")
        thin_border   = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        center_align  = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center_align
            cell.border    = thin_border

        for row_idx in range(2, ws.max_row + 1):
            fill = alt_fill if row_idx % 2 == 0 else PatternFill()
            for cell in ws[row_idx]:
                cell.fill   = fill
                cell.border = thin_border

        for col_idx in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col_idx)].auto_size = True
            width = max(
                len(str(ws.cell(row=r, column=col_idx).value or ""))
                for r in range(1, ws.max_row + 1)
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(width + 4, 60)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        wb.save(path)

    @staticmethod
    def to_html(components: list[Component], path: str, project_name: str = "BOM") -> None:
        rows = BOMGenerator.group_components(components)
        if not rows:
            Path(path).write_text("<p>Brak komponentów</p>", encoding="utf-8")
            return

        total_qty = sum(r["Ilość"] for r in rows)
        type_counts: dict = {}
        for r in rows:
            type_counts[r["Typ"]] = type_counts.get(r["Typ"], 0) + r["Ilość"]
        summary = " | ".join(f"<b>{v}</b>× {k}" for k, v in sorted(type_counts.items(), key=lambda x: -x[1]))

        columns = list(rows[0].keys())
        header_html = "".join(f"<th>{col}</th>" for col in columns)
        rows_html = ""
        for i, row in enumerate(rows):
            css = "odd" if i % 2 == 0 else "even"
            cells = "".join(
                f'<td><a href="{row["Datasheet"]}" target="_blank">{row["Datasheet"][:40]}…</a></td>'
                if col == "Datasheet" and row.get("Datasheet")
                else f"<td>{row.get(col,'')}</td>"
                for col in columns
            )
            rows_html += f'<tr class="{css}">{cells}</tr>\n'

        html = f"""<!DOCTYPE html>
<html lang="pl">
<head>
<meta charset="UTF-8">
<title>BOM — {project_name}</title>
<style>
  body {{ font-family: 'Segoe UI', Arial, sans-serif; background:#1a1a2e; color:#e0e0e0; margin:20px; }}
  h1 {{ color:#4fc3f7; border-bottom:2px solid #333; padding-bottom:8px; }}
  .summary {{ background:#0d2137; border-left:4px solid #4fc3f7; padding:10px 16px; margin:12px 0; border-radius:4px; }}
  table {{ border-collapse:collapse; width:100%; margin-top:16px; font-size:13px; }}
  th {{ background:#0d47a1; color:#fff; padding:10px 12px; text-align:left; white-space:nowrap; }}
  tr.odd  {{ background:#1a1a2e; }}
  tr.even {{ background:#0a1929; }}
  tr:hover {{ background:#1565c0 !important; cursor:pointer; }}
  td {{ padding:7px 12px; border-bottom:1px solid #333; }}
  a {{ color:#4fc3f7; text-decoration:none; }}
  a:hover {{ text-decoration:underline; }}
  .badge {{ display:inline-block; background:#0d47a1; color:#fff; border-radius:12px;
             padding:2px 8px; font-size:11px; margin-right:4px; }}
  @media print {{ body {{ background:#fff; color:#000; }} th {{ background:#1565c0; }} }}
</style>
</head>
<body>
<h1>📋 BOM — {project_name}</h1>
<div class="summary">
  Łącznie: <b>{len(rows)}</b> pozycji, <b>{total_qty}</b> elementów &nbsp;|&nbsp; {summary}
</div>
<table>
  <thead><tr>{header_html}</tr></thead>
  <tbody>{rows_html}</tbody>
</table>
<p style="color:#555;font-size:11px;margin-top:20px;">
  Wygenerowano przez ElectroVision &nbsp;|&nbsp; {project_name}
</p>
</body>
</html>"""
        Path(path).write_text(html, encoding="utf-8")
